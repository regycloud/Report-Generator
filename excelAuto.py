from pickle import FALSE
from unittest import registerResult
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.styles import numbers
from percentileFinder import findValue


# ensure the filename, sequence, and cid.

startPicture = 1
endPicture = 1
endOfMonth = True


wb = load_workbook(
    r'014_FT2107_01_Daily Burst Traffic Utilization Report_SSPL.VAL.13.02_July 2021_PRTG.xlsx')
cidOptionOne = ['sspl.mix1.011', -64, [375, 658], [1224, 583], 8290]
cidOptionTwo = ['grna.val.73.01', -70, [372, 612], [1224, 537], 9025]
cidOptionThree = ['sspl.val.73.02', -70, [375, 611], [1224, 537], 9028]
cidOptionFour = ['sspl.val.73.01', -70, [375, 611], [1224, 537], 8982]
cidOptionFive = ['bke.val.73.01', -70, [375, 611], [1224, 537], 9091]
cidOptioneSix = ['0110306', -70, [375, 611],
                 [1224, 537], 8833]  # SSPL.VAL.13.02 = Link#1
cidOptionSeven = ['0210304', -70, [375, 611],
                  [1224, 537], 8834]  # SSPL.VAL.13.03 = Link#2

# CID key: 
#  {'link1': ['77,565', '117,942'], 'link2': ['82,562', '117,960'], 'MIX1': ['5', '65,328'], 'SSPLVAL7302': ['17,349', '7,173'], 'SSPLVAL7301': ['24,471', '49,070'], 'BKEVAL7301': ['1,917', '3,609']}
your_cid = 'MIX1'

print('This will be working on row', startPicture, 'until', endPicture)
cont = input(
    'Please ensure your CID is {} , and please check the filename. Are those okay? y/n '.format(your_cid))
if cont == 'n':
    print('Please change on the your cid and/or file.')
    exit()
ws = wb.active


p2e = pixels_to_EMU


position = XDRPoint2D(p2e(500), p2e(500))
size = XDRPositiveSize2D(p2e(307.2), p2e(86.4))
size2 = XDRPositiveSize2D(p2e(310.08), p2e(25.92))

c2e = cm_to_EMU

# Calculated number of cells width or height from cm into EMUs


def cellh(x): return c2e((x * 49.77)/99)
def cellw(x): return c2e((x * (18.65-1.71))/10)

# Want to place image in row 5 (6 in excel), column 2 (C in excel)
# Also offset by half a column.


column = 10
coloffset = cellw(0.3)
row = 36
rowoffset = cellh(0.7)

for i in range(startPicture, endPicture + 1):
    # This is for graph
    img = Image('{} - SSPL.MIX.011_.png'.format(i))
    # This is for table
    img2 = Image('{} - SSPL.MIX.011_.png'.format(i))

    try:
        if your_cid:
            result = findValue('{}'.format(i))
            result1 = result[0][0].replace('.',',')
            result2 = result[0][1].replace('.',',')
            cellRow = 36 + i
            print(result1, result2, 'cell row', cellRow)
            ws['D{}'.format(cellRow)] = result1
            ws['F{}'.format(cellRow)]  = result2  
            ws['D{}'.format(cellRow)].number_format = numbers.FORMAT_NUMBER
            ws['F{}'.format(cellRow)].number_format = numbers.FORMAT_NUMBER
        # insertValue = findValue('{}.1'.format(i))
        # ws['D{}'.format(row + i)] = insertValue[0]
        # ws['F{}'.format(row + i)] = insertValue[1]

    except:
        print('data unavailable')

    h, w = img.height, img.width

    if i != 1:
        row = row + 1

    marker1 = AnchorMarker(col=column, colOff=coloffset,
                           row=row, rowOff=rowoffset)

    marker2 = AnchorMarker(col=column, colOff=coloffset,
                           row=row, rowOff=rowoffset + cellh(5.5))

    img.anchor = OneCellAnchor(_from=marker1, ext=size)
    img2.anchor = OneCellAnchor(_from=marker2, ext=size2)
    ws.add_image(img)
    ws.add_image(img2)
    print('working for {} table'.format(i))

if endOfMonth == True:

    try:
        if your_cid:
            result = findValue('32')
            # Average field

            result1 = result[1][0].replace('.',',')
            result2 = result[1][1].replace('.',',')
            cellRow = 69
            print(result1, result2, 'cell row', cellRow)
            ws['D{}'.format(cellRow)] = result1
            ws['F{}'.format(cellRow)]  = result2  
            ws['D{}'.format(cellRow)].number_format = numbers.FORMAT_NUMBER
            ws['F{}'.format(cellRow)].number_format = numbers.FORMAT_NUMBER
            
            result1 = result[0][0].replace('.',',')
            result2 = result[0][1].replace('.',',')
            cellRow = 71
            print(result1, result2, 'cell row', cellRow)
            ws['D{}'.format(cellRow)] = result1
            ws['F{}'.format(cellRow)]  = result2  
            ws['D{}'.format(cellRow)].number_format = numbers.FORMAT_NUMBER
            ws['F{}'.format(cellRow)].number_format = numbers.FORMAT_NUMBER

    except:
        print('data not found')

    img = Image('32 - SSPL.MIX.011_.png'.format(i))
    # This is for table
    img2 = Image('32 - SSPL.MIX.011_.png'.format(i))
    marker1 = AnchorMarker(col=10, colOff=coloffset, row=70, rowOff=rowoffset)

    marker2 = AnchorMarker(col=10, colOff=coloffset,
                           row=70, rowOff=rowoffset + cellh(5.5))

    img.anchor = OneCellAnchor(_from=marker1, ext=size)
    img2.anchor = OneCellAnchor(_from=marker2, ext=size2)
    ws.add_image(img)
    ws.add_image(img2)
wb.save('result.xlsx')


# done all
