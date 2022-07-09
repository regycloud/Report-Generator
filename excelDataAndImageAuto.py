from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.styles import numbers
from percentileFinder import findValue


# ensure the filename, sequence, and cid.
def createReport(cid, startPicture, endPicture):
    endOfMonth = True
    fileName = '*.xlsx'

    # Load the workbook
    wb = load_workbook(fileName)


    cid = 1

    if (cid == 0):
        newNameFile = '{} - SSPL.VAL.13.02.png'.format('date')
    if (cid == 1):
        newNameFile = '{} - SSPL.VAL.13.03.png'.format('date')
    if (cid == 2):
        newNameFile = '{} - SSPL.MIX.011_.png'.format('date')
    if (cid == 3):
        newNameFile = '{} - SSPL.VAL.73.02.png'.format('date')
    if (cid == 4):
        newNameFile = '{} - SSPL.VAL.73.01.png'.format('date')
    if (cid == 5):
        newNameFile = '{} - BKE.VAL.73.01.png'.format('date')
    if (cid == 6):
        newNameFile = '{} - GRNA.VAL.73.01.png'.format('date')


    print('This will be working on row', startPicture, 'until', endPicture)
    cont = input(
        'Please ensure your CID is {} , and please check the filename. Are those okay? y/n '.format(cid))
    if cont == 'n':
        print('Please change on the your cid and/or file.')
        exit()

    ws = wb.active


    p2e = pixels_to_EMU


    position = XDRPoint2D(p2e(500), p2e(500))
    size = XDRPositiveSize2D(p2e(307.2), p2e(86.4))
    size2 = XDRPositiveSize2D(p2e(310.08), p2e(25.92))

    c2e = cm_to_EMU

    # Calculated number of cells width or height from cm into EMUs


    def cellh(x): return c2e((x * 49.77)/99)
    def cellw(x): return c2e((x * (18.65-1.71))/10)

    # Want to place image in row 5 (6 in excel), column 2 (C in excel)
    # Also offset by half a column.


    column = 10
    coloffset = cellw(0.3)
    row = 36
    rowoffset = cellh(0.7)

    for i in range(startPicture, endPicture):
        # This is for graph
        if (cid == 0):
            newNameFile = '{} - SSPL.VAL.13.02.png'.format(i)
        if (cid == 1):
            newNameFile = '{} - SSPL.VAL.13.03.png'.format(i)
        if (cid == 2):
            newNameFile = '{} - SSPL.MIX.011_.png'.format(i)
        if (cid == 3):
            newNameFile = '{} - SSPL.VAL.73.02.png'.format(i)
        if (cid == 4):
            newNameFile = '{} - SSPL.VAL.73.01.png'.format(i)
        if (cid == 5):
            newNameFile = '{} - BKE.VAL.73.01.png'.format(i)
        if (cid == 6):
            newNameFile = '{} - GRNA.VAL.73.01.png'.format(i)

        img = Image('./imgs/graph/{}'.format(newNameFile))
        # This is for table
        imgTable = Image('./imgs/table/{}'.format(newNameFile))

        try:
            if newNameFile:
                result = findValue('./imgs/table/{}'.format(newNameFile))
                result1 = result[0][0].replace('.',',')
                result2 = result[0][1].replace('.',',')
                cellRow = 36 + i
                print(result1, result2, 'cell row', cellRow)
                ws['D{}'.format(cellRow)] = result1
                ws['F{}'.format(cellRow)]  = result2  
                ws['D{}'.format(cellRow)].number_format = numbers.FORMAT_NUMBER
                ws['F{}'.format(cellRow)].number_format = numbers.FORMAT_NUMBER
            # insertValue = findValue('{}.1'.format(i))
            # ws['D{}'.format(row + i)] = insertValue[0]
            # ws['F{}'.format(row + i)] = insertValue[1]

        except:
            print('data unavailable')

        h, w = img.height, img.width

        if i != 1:
            row = row + 1

        marker1 = AnchorMarker(col=column, colOff=coloffset,
                            row=row, rowOff=rowoffset)

        marker2 = AnchorMarker(col=column, colOff=coloffset,
                            row=row, rowOff=rowoffset + cellh(5.5))

        img.anchor = OneCellAnchor(_from=marker1, ext=size)
        imgTable.anchor = OneCellAnchor(_from=marker2, ext=size2)
        ws.add_image(img)
        ws.add_image(imgTable)
        print('working for {} table'.format(i))

    if endOfMonth == True:

        try:
            if newNameFile:
                newNameFile = newNameFile[2:]
                result = findValue('./imgs/table/32{}'.format(newNameFile))
                # Average field

                result1 = result[1][0].replace('.',',')
                result2 = result[1][1].replace('.',',')
                cellRow = 69
                print(result1, result2, 'cell row', cellRow)
                ws['D{}'.format(cellRow)] = result1
                ws['F{}'.format(cellRow)]  = result2  
                ws['D{}'.format(cellRow)].number_format = numbers.FORMAT_NUMBER
                ws['F{}'.format(cellRow)].number_format = numbers.FORMAT_NUMBER
                
                result1 = result[0][0].replace('.',',')
                result2 = result[0][1].replace('.',',')
                cellRow = 71
                print(result1, result2, 'cell row', cellRow)
                ws['D{}'.format(cellRow)] = result1
                ws['F{}'.format(cellRow)]  = result2  
                ws['D{}'.format(cellRow)].number_format = numbers.FORMAT_NUMBER
                ws['F{}'.format(cellRow)].number_format = numbers.FORMAT_NUMBER

        except:
            print('data not found')

        img = Image('./imgs/graph/32{}'.format(newNameFile))
        # This is for table
        imgTable = Image('./imgs/table/32{}'.format(newNameFile))
        marker1 = AnchorMarker(col=10, colOff=coloffset, row=70, rowOff=rowoffset)

        marker2 = AnchorMarker(col=10, colOff=coloffset,
                            row=70, rowOff=rowoffset + cellh(5.5))

        img.anchor = OneCellAnchor(_from=marker1, ext=size)
        imgTable.anchor = OneCellAnchor(_from=marker2, ext=size2)
        ws.add_image(img)
        ws.add_image(imgTable)
    wb.save('[completed]{}.xlsx'.format(fileName))

createReport(0, 1, 32)


